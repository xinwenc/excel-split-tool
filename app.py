import json
import os
import re
import shutil
import uuid
import zipfile
from copy import copy
from datetime import date, datetime

from flask import Flask, after_this_request, jsonify, render_template, request, send_file
from openpyxl import Workbook, load_workbook
from werkzeug.utils import secure_filename

app = Flask(__name__)

UPLOAD_FOLDER = "uploads"
ALLOWED_EXTENSIONS = {"xlsx", "xlsm"}
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
app.config["MAX_CONTENT_LENGTH"] = 100 * 1024 * 1024  # 100MB
os.makedirs(UPLOAD_FOLDER, exist_ok=True)


# ----------------------------- 基础工具函数 -----------------------------
def allowed_file(filename: str) -> bool:
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def save_temp_upload(file_storage) -> str:
    """保存上传文件到临时路径。"""
    filename = secure_filename(file_storage.filename)
    temp_path = os.path.join(app.config["UPLOAD_FOLDER"], f"temp_{uuid.uuid4()}_{filename}")
    file_storage.save(temp_path)
    return temp_path


def cleanup_path(path: str):
    """安全删除文件或文件夹。"""
    try:
        if os.path.isdir(path):
            shutil.rmtree(path)
        elif os.path.exists(path):
            os.remove(path)
    except Exception:
        pass


def make_value_key(value) -> str:
    """统一拆分值格式，保证前端获取值和后端拆分时比较一致。"""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def safe_filename_part(value) -> str:
    """把拆分值转换为安全的文件名片段。"""
    text = make_value_key(value)
    text = re.sub(r"[\\/*?:\"<>|]", "_", text)
    text = text.strip().strip(".")
    return text or "空值"


def unique_output_path(output_dir: str, filename: str) -> str:
    """如果文件名重复，自动追加序号，避免覆盖。"""
    base, ext = os.path.splitext(filename)
    candidate = os.path.join(output_dir, filename)
    counter = 2
    while os.path.exists(candidate):
        candidate = os.path.join(output_dir, f"{base}_{counter}{ext}")
        counter += 1
    return candidate


def get_merged_cell_value(ws, row: int, col: int):
    """读取单元格值；如果该格属于合并单元格，则读取合并区域左上角的值。"""
    cell = ws.cell(row, col)
    if cell.coordinate in ws.merged_cells:
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                return ws.cell(merged_range.min_row, merged_range.min_col).value
    return cell.value


def get_sheet_names(file_path: str):
    wb = load_workbook(file_path, data_only=True, read_only=False)
    sheets = wb.sheetnames
    wb.close()
    return sheets


def get_columns_from_sheet(file_path: str, sheet_name: str, header_row: int):
    """读取指定工作表的表头行，返回列号+列名。支持合并单元格。"""
    wb = load_workbook(file_path, data_only=True, read_only=False)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"工作表不存在：{sheet_name}")

        ws = wb[sheet_name]
        if header_row < 1 or header_row > ws.max_row:
            raise ValueError(f"表头行号超出范围：{header_row}")

        columns = []
        for col in range(1, ws.max_column + 1):
            value = get_merged_cell_value(ws, header_row, col)
            name = make_value_key(value) or f"列{col}"
            columns.append({"index": col, "name": name, "label": f"第{col}列：{name}"})
        return columns
    finally:
        wb.close()


def parse_sheet_configs(raw_value: str):
    """解析前端传来的每个拆分 sheet 的配置。"""
    try:
        configs = json.loads(raw_value or "[]")
    except Exception:
        raise ValueError("sheet_configs 格式错误")

    if not isinstance(configs, list) or not configs:
        raise ValueError("请至少配置一个拆分工作表")

    normalized = []
    seen = set()
    for item in configs:
        if not isinstance(item, dict):
            raise ValueError("sheet_configs 中存在非法配置")

        sheet_name = make_value_key(item.get("sheet_name"))
        if not sheet_name:
            raise ValueError("存在未填写工作表名称的配置")
        if sheet_name in seen:
            raise ValueError(f"拆分工作表重复：{sheet_name}")
        seen.add(sheet_name)

        try:
            header_row = int(item.get("header_row", 1))
        except Exception:
            raise ValueError(f"工作表【{sheet_name}】的表头行号必须是数字")
        if header_row < 1:
            raise ValueError(f"工作表【{sheet_name}】的表头行号不能小于 1")

        split_mode = item.get("split_mode", "col_name")
        if split_mode not in {"col_name", "col_index"}:
            raise ValueError(f"工作表【{sheet_name}】的拆分方式不正确")

        normalized.append({
            "sheet_name": sheet_name,
            "header_row": header_row,
            "split_mode": split_mode,
            "split_value": make_value_key(item.get("split_value")),
            "split_col_name": make_value_key(item.get("split_col_name")),
            "split_col_index": item.get("split_col_index"),
        })

    return normalized


def resolve_split_column(ws, config: dict) -> int:
    """根据单个 sheet 的配置确定拆分列号。"""
    sheet_name = config["sheet_name"]
    header_row = config["header_row"]
    split_mode = config["split_mode"]

    # 前端按列名选择时，也会把实际列号传回来，优先使用列号，避免重名表头导致找错。
    split_col_index = config.get("split_col_index")
    if split_col_index not in (None, ""):
        try:
            col_num = int(split_col_index)
        except Exception:
            raise ValueError(f"工作表【{sheet_name}】的拆分列号无效")
        if col_num < 1 or col_num > ws.max_column:
            raise ValueError(f"工作表【{sheet_name}】的拆分列号超出范围：{col_num}")
        return col_num

    if split_mode == "col_index":
        try:
            col_num = int(config.get("split_value"))
        except Exception:
            raise ValueError(f"工作表【{sheet_name}】的拆分列号必须是数字")
        if col_num < 1 or col_num > ws.max_column:
            raise ValueError(f"工作表【{sheet_name}】的拆分列号超出范围：{col_num}")
        return col_num

    # 按列名查找
    target_name = config.get("split_col_name") or config.get("split_value")
    target_name = make_value_key(target_name)
    if not target_name:
        raise ValueError(f"工作表【{sheet_name}】请选择拆分列名")

    for col in range(1, ws.max_column + 1):
        cell_name = make_value_key(get_merged_cell_value(ws, header_row, col)) or f"列{col}"
        if cell_name == target_name:
            return col

    raise ValueError(f"工作表【{sheet_name}】未找到拆分列：{target_name}")


# ----------------------------- 样式复制函数：完整模式优化版 -----------------------------
def copy_cell_with_style(src_cell, dst_cell, style_cache=None):
    """
    复制单元格值和完整静态样式。

    优化点：
    旧版每个单元格都重新 copy(font/border/fill/alignment/protection)，很慢。
    新版按源单元格 style_id 做缓存：同一种样式只复制一次，后续单元格复用同一组样式对象。
    这样比直接复制 _style 更安全，不会产生目标工作簿样式 ID 不匹配的问题。

    当前工作簿仍使用 data_only=True 读取，所以公式会复制为 Excel 已缓存的计算结果。
    """
    dst_cell.value = src_cell.value

    if src_cell.has_style:
        cache = style_cache if style_cache is not None else {}
        style_id = src_cell.style_id
        cached_style = cache.get(style_id)
        if cached_style is None:
            cached_style = {
                "font": copy(src_cell.font),
                "border": copy(src_cell.border),
                "fill": copy(src_cell.fill),
                "number_format": src_cell.number_format,
                "alignment": copy(src_cell.alignment),
                "protection": copy(src_cell.protection),
            }
            cache[style_id] = cached_style

        dst_cell.font = cached_style["font"]
        dst_cell.border = cached_style["border"]
        dst_cell.fill = cached_style["fill"]
        dst_cell.number_format = cached_style["number_format"]
        dst_cell.alignment = cached_style["alignment"]
        dst_cell.protection = cached_style["protection"]

    if src_cell.hyperlink:
        dst_cell._hyperlink = copy(src_cell.hyperlink)
    if src_cell.comment:
        dst_cell.comment = copy(src_cell.comment)

def copy_column_dimensions(src_ws, dst_ws):
    for col_letter, dim in src_ws.column_dimensions.items():
        dst_dim = dst_ws.column_dimensions[col_letter]
        dst_dim.width = dim.width
        dst_dim.hidden = dim.hidden
        dst_dim.outlineLevel = dim.outlineLevel
        dst_dim.collapsed = dim.collapsed
        dst_dim.bestFit = dim.bestFit


def copy_row_dimension(src_ws, dst_ws, src_row: int, dst_row: int):
    # 只复制原表里显式设置过的行高/隐藏/分组，避免为每一行创建无意义的 RowDimension，提高速度并减小文件体积。
    if src_row not in src_ws.row_dimensions:
        return
    src_dim = src_ws.row_dimensions[src_row]
    dst_dim = dst_ws.row_dimensions[dst_row]
    dst_dim.height = src_dim.height
    dst_dim.hidden = src_dim.hidden
    dst_dim.outlineLevel = src_dim.outlineLevel
    dst_dim.collapsed = src_dim.collapsed
    dst_dim.ht = src_dim.ht


def copy_basic_sheet_settings(src_ws, dst_ws):
    copy_column_dimensions(src_ws, dst_ws)
    dst_ws.freeze_panes = src_ws.freeze_panes
    try:
        dst_ws.sheet_format = copy(src_ws.sheet_format)
        dst_ws.sheet_properties = copy(src_ws.sheet_properties)
        dst_ws.page_margins = copy(src_ws.page_margins)
        dst_ws.page_setup = copy(src_ws.page_setup)
        dst_ws.print_options = copy(src_ws.print_options)
        dst_ws.sheet_view.showGridLines = src_ws.sheet_view.showGridLines
    except Exception:
        pass


def copy_header_merged_cells(src_ws, dst_ws, header_row: int):
    """只复制完全位于表头区域内的合并单元格，避免数据行筛选后合并区域错位。"""
    for merged_range in src_ws.merged_cells.ranges:
        if merged_range.max_row <= header_row:
            dst_ws.merge_cells(str(merged_range))


def copy_all_merged_cells(src_ws, dst_ws):
    for merged_range in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merged_range))


def collect_conditional_formatting_rules(src_ws):
    """
    提前收集条件格式规则。

    这样在同一个拆分任务内，不用每生成一个文件都重新遍历源工作表的 conditional_formatting 容器；
    复制到目标 sheet 时仍然会 copy(rule)，保证每个新 workbook 有独立规则对象。
    """
    rules = []
    try:
        for conditional_format in src_ws.conditional_formatting:
            sqref = str(conditional_format.sqref)
            rules.append((sqref, list(conditional_format.rules)))
    except Exception:
        pass
    return rules


def apply_conditional_formatting(dst_ws, cached_rules):
    """把提前缓存的条件格式规则应用到目标工作表。"""
    if not cached_rules:
        return
    try:
        for sqref, rules in cached_rules:
            for rule in rules:
                dst_ws.conditional_formatting.add(sqref, copy(rule))
    except Exception:
        # 条件格式复制失败不影响主流程，至少保证文件可以正常生成。
        pass


def copy_cells_rect(src_ws, dst_ws, src_row: int, dst_row: int, max_col: int, style_cache=None):
    """复制一整行指定列范围内的单元格。"""
    copy_row_dimension(src_ws, dst_ws, src_row, dst_row)
    for col in range(1, max_col + 1):
        src_cell = src_ws.cell(src_row, col)
        # 空白且无样式/批注/超链接的单元格没有必要写入，能减少大量无效对象创建。
        if src_cell.value is None and not src_cell.has_style and not src_cell.comment and not src_cell.hyperlink:
            continue
        copy_cell_with_style(src_cell, dst_ws.cell(dst_row, col), style_cache=style_cache)


def copy_full_sheet(src_ws, dst_ws, cached_cf_rules=None):
    """完整复制保留工作表。"""
    copy_basic_sheet_settings(src_ws, dst_ws)
    max_col = src_ws.max_column
    style_cache = {}
    for row in range(1, src_ws.max_row + 1):
        copy_cells_rect(src_ws, dst_ws, row, row, max_col, style_cache=style_cache)
    copy_all_merged_cells(src_ws, dst_ws)
    apply_conditional_formatting(dst_ws, cached_cf_rules if cached_cf_rules is not None else collect_conditional_formatting_rules(src_ws))


def copy_split_sheet_by_rows(src_ws, dst_ws, header_row: int, matched_rows: list, cached_cf_rules=None):
    """
    复制单个拆分工作表：表头完整保留，数据只复制提前索引好的行。

    优化点：
    旧版每生成一个拆分值，就从 header_row+1 扫到 max_row 判断一次。
    新版在生成前已经建立“拆分值 -> 源行号列表”的索引，这里直接复制 matched_rows。
    """
    copy_basic_sheet_settings(src_ws, dst_ws)
    max_col = src_ws.max_column

    # 复制表头及表头上方
    style_cache = {}
    for row in range(1, header_row + 1):
        copy_cells_rect(src_ws, dst_ws, row, row, max_col, style_cache=style_cache)
    copy_header_merged_cells(src_ws, dst_ws, header_row)

    # 复制匹配数据行
    target_row = header_row + 1
    for src_row in matched_rows:
        copy_cells_rect(src_ws, dst_ws, src_row, target_row, max_col, style_cache=style_cache)
        target_row += 1

    # 完整模式：继续复制条件格式，保留 AV 列这类动态颜色。
    apply_conditional_formatting(dst_ws, cached_cf_rules if cached_cf_rules is not None else collect_conditional_formatting_rules(src_ws))

    # 设置筛选区域，方便打开结果后查看
    if dst_ws.max_row >= header_row and dst_ws.max_column >= 1:
        try:
            dst_ws.auto_filter.ref = dst_ws.dimensions
        except Exception:
            pass


# ----------------------------- 核心拆分逻辑：一次建索引，避免重复扫描 -----------------------------
def prepare_sheet_configs(wb, sheet_configs: list):
    """预先校验每个拆分 sheet，并解析拆分列号、条件格式缓存。"""
    prepared_configs = []
    split_sheet_names = set()

    for config in sheet_configs:
        sheet_name = config["sheet_name"]
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"工作表不存在：{sheet_name}")
        ws = wb[sheet_name]
        if config["header_row"] < 1 or config["header_row"] > ws.max_row:
            raise ValueError(f"工作表【{sheet_name}】的表头行号超出范围")
        split_col = resolve_split_column(ws, config)
        prepared_configs.append({
            **config,
            "split_col": split_col,
            "cf_rules": collect_conditional_formatting_rules(ws),
            "rows_by_value": {},
        })
        split_sheet_names.add(sheet_name)

    return prepared_configs, split_sheet_names


def build_rows_index(wb, prepared_configs: list, selected_values=None):
    """
    一次性扫描所有拆分工作表，建立索引：
    每个 sheet 内部：拆分值 -> 对应源行号列表。

    selected_values 不为空时，只为用户勾选的拆分值建立行索引，减少内存和后续处理量。
    """
    selected_set = None
    if selected_values is not None:
        selected_set = {make_value_key(v) for v in selected_values if make_value_key(v)}

    all_values = set()

    for config in prepared_configs:
        ws = wb[config["sheet_name"]]
        split_col = config["split_col"]
        header_row = config["header_row"]
        rows_by_value = {}

        if header_row < ws.max_row:
            for row in range(header_row + 1, ws.max_row + 1):
                value = make_value_key(ws.cell(row, split_col).value)
                if not value:
                    continue
                all_values.add(value)
                if selected_set is None or value in selected_set:
                    rows_by_value.setdefault(value, []).append(row)

        config["rows_by_value"] = rows_by_value

    return sorted(all_values)


def collect_split_values(file_path: str, sheet_configs: list):
    """从所有已配置的拆分 sheet 中汇总拆分值。"""
    wb = load_workbook(file_path, data_only=True, read_only=False)
    try:
        prepared_configs, _ = prepare_sheet_configs(wb, sheet_configs)
        return build_rows_index(wb, prepared_configs, selected_values=None)
    finally:
        wb.close()


def split_excel_by_sheet_configs(original_file_path: str, output_dir: str, sheet_configs: list,
                                  keep_sheets: list, filename_template: str, selected_values: list):
    """
    按多个 sheet 的独立配置进行拆分。
    每个拆分值生成一个 Excel；每个 Excel 内包含所有被选中的拆分 sheet，且每个 sheet 只保留该拆分值的数据。

    优化版核心：
    1）先解析所有 sheet 的拆分列；
    2）一次性扫描所有 sheet，建立“拆分值 -> 行号列表”索引；
    3）生成每个 Excel 时直接按行号复制，不再重复全表扫描；
    4）完整模式仍保留静态样式、条件格式、合并单元格、列宽、行高等。
    """
    wb = load_workbook(original_file_path, data_only=True, read_only=False)
    base_name = os.path.splitext(os.path.basename(original_file_path))[0]

    try:
        prepared_configs, split_sheet_names = prepare_sheet_configs(wb, sheet_configs)

        values_to_split = [make_value_key(v) for v in selected_values if make_value_key(v)]
        # 保持用户勾选顺序，同时去重
        values_to_split = list(dict.fromkeys(values_to_split))
        if not values_to_split:
            return False, "没有选择任何拆分值"

        # 关键优化：只扫描一次所有 sheet，建立索引。
        build_rows_index(wb, prepared_configs, selected_values=values_to_split)

        # 预先缓存完整保留工作表的条件格式，避免每个文件都重新遍历源 sheet 条件格式容器。
        keep_sheet_cf_cache = {}
        for sheet_name in keep_sheets:
            if sheet_name in wb.sheetnames and sheet_name not in split_sheet_names:
                keep_sheet_cf_cache[sheet_name] = collect_conditional_formatting_rules(wb[sheet_name])

        for value in values_to_split:
            safe_value = safe_filename_part(value)
            filename = filename_template.replace("{value}", safe_value).replace("{base_name}", base_name)
            filename = re.sub(r"[\\/*?:\"<>|]", "_", filename).strip()
            if not filename:
                filename = f"{safe_value}_{base_name}"
            if not filename.lower().endswith(".xlsx"):
                filename += ".xlsx"
            new_filepath = unique_output_path(output_dir, filename)

            new_wb = Workbook()
            new_wb.remove(new_wb.active)

            # 拆分工作表：每个 sheet 使用自己的表头行和拆分列；数据行直接按索引复制。
            for config in prepared_configs:
                sheet_name = config["sheet_name"]
                old_ws = wb[sheet_name]
                new_ws = new_wb.create_sheet(title=sheet_name)
                matched_rows = config.get("rows_by_value", {}).get(value, [])
                copy_split_sheet_by_rows(
                    old_ws,
                    new_ws,
                    header_row=config["header_row"],
                    matched_rows=matched_rows,
                    cached_cf_rules=config.get("cf_rules"),
                )

            # 完整保留工作表：如果已经作为拆分 sheet，则不重复创建。
            for sheet_name in keep_sheets:
                if sheet_name not in wb.sheetnames:
                    continue
                if sheet_name in split_sheet_names:
                    continue
                old_ws = wb[sheet_name]
                new_ws = new_wb.create_sheet(title=sheet_name)
                copy_full_sheet(old_ws, new_ws, cached_cf_rules=keep_sheet_cf_cache.get(sheet_name))

            new_wb.save(new_filepath)

        return True, f"成功拆分为 {len(values_to_split)} 个文件"
    except Exception as e:
        return False, str(e)
    finally:
        wb.close()


# ----------------------------- Flask 路由 -----------------------------
@app.route("/")
def index():
    return render_template("index.html")


@app.route("/get_sheets_and_columns", methods=["POST"])
def get_sheets_and_columns():
    if "file" not in request.files:
        return jsonify({"error": "没有文件"}), 400
    file = request.files["file"]
    if file.filename == "":
        return jsonify({"error": "文件名为空"}), 400
    if not allowed_file(file.filename):
        return jsonify({"error": "仅支持 .xlsx / .xlsm 文件，请将 .xls 另存为 .xlsx 后再上传"}), 400

    temp_path = save_temp_upload(file)
    try:
        sheet_name = request.form.get("sheet_name")
        header_row = request.form.get("header_row")
        if sheet_name and header_row:
            columns = get_columns_from_sheet(temp_path, sheet_name, int(header_row))
            result = {"columns": columns}
        else:
            sheets = get_sheet_names(temp_path)
            result = {"sheets": sheets}
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        cleanup_path(temp_path)


@app.route("/get_split_values", methods=["POST"])
def get_split_values():
    if "file" not in request.files:
        return jsonify({"error": "没有文件"}), 400
    file = request.files["file"]
    if file.filename == "":
        return jsonify({"error": "文件名为空"}), 400
    if not allowed_file(file.filename):
        return jsonify({"error": "仅支持 .xlsx / .xlsm 文件，请将 .xls 另存为 .xlsx 后再上传"}), 400

    temp_path = save_temp_upload(file)
    try:
        sheet_configs = parse_sheet_configs(request.form.get("sheet_configs", "[]"))
        values = collect_split_values(temp_path, sheet_configs)
        return jsonify({"values": values})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        cleanup_path(temp_path)


@app.route("/split", methods=["POST"])
def split_file():
    if "file" not in request.files:
        return "没有文件", 400
    file = request.files["file"]
    if file.filename == "":
        return "文件名为空", 400
    if not allowed_file(file.filename):
        return "仅支持 .xlsx / .xlsm 文件，请将 .xls 另存为 .xlsx 后再上传", 400

    filename_template = request.form.get("filename_template", "{value}_{base_name}").strip() or "{value}_{base_name}"
    selected_values = request.form.getlist("selected_values")
    keep_sheets = request.form.getlist("keep_sheets")

    try:
        sheet_configs = parse_sheet_configs(request.form.get("sheet_configs", "[]"))
    except Exception as e:
        return f"拆分失败：{str(e)}", 400

    if not selected_values:
        return "请至少选择一个要拆分的值", 400

    filename = secure_filename(file.filename)
    input_path = os.path.join(app.config["UPLOAD_FOLDER"], f"input_{uuid.uuid4()}_{filename}")
    file.save(input_path)

    unique_id = str(uuid.uuid4())
    output_dir = os.path.join(app.config["UPLOAD_FOLDER"], f"split_{unique_id}")
    os.makedirs(output_dir, exist_ok=True)

    try:
        success, message = split_excel_by_sheet_configs(
            input_path,
            output_dir,
            sheet_configs,
            keep_sheets,
            filename_template,
            selected_values,
        )

        if not success:
            return f"拆分失败：{message}", 400

        zip_name = f"{os.path.splitext(filename)[0]}_拆分结果.zip"
        zip_path = os.path.join(app.config["UPLOAD_FOLDER"], zip_name)

        # compresslevel=1 比默认压缩更快；仍然是 zip 压缩包，只是优先速度。
        try:
            zip_file = zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED, compresslevel=1)
        except TypeError:
            # 兼容较旧 Python 版本。
            zip_file = zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED)

        with zip_file as zipf:
            for root, _, files in os.walk(output_dir):
                for f in files:
                    full_path = os.path.join(root, f)
                    zipf.write(full_path, arcname=f)

        @after_this_request
        def remove_zip(response):
            cleanup_path(zip_path)
            return response

        return send_file(zip_path, as_attachment=True, download_name=zip_name)
    finally:
        cleanup_path(input_path)
        cleanup_path(output_dir)


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(debug=True, host="0.0.0.0", port=port)
